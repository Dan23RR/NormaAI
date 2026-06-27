"""
Importa i risultati JSON del test runner nella Confusion Matrix Excel.

Uso:
    python results_to_excel.py [test_results.json]
"""
import json
import os
import sys

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

results_file = sys.argv[1] if len(sys.argv) > 1 else "test_results.json"
xlsx_file = os.path.join(os.path.dirname(__file__), "NormaAI_Confusion_Matrix.xlsx")

with open(os.path.join(os.path.dirname(__file__), results_file)) as f:
    data = json.load(f)

tp_fill = PatternFill("solid", fgColor="C6EFCE")
pm_fill = PatternFill("solid", fgColor="FFEB9C")
fn_fill = PatternFill("solid", fgColor="FFC7CE")
err_fill = PatternFill("solid", fgColor="E0E0E0")
fill_map = {"TRUE_POSITIVE": tp_fill, "PARTIAL_MATCH": pm_fill, "FALSE_NEGATIVE": fn_fill, "ERROR": err_fill}

wb = load_workbook(xlsx_file)
ws = wb["Confusion Matrix"]

# Build test_id -> row mapping
id_to_row = {}
for row in range(5, ws.max_row + 1):
    tid = ws.cell(row=row, column=1).value
    if tid:
        id_to_row[tid] = row

updated = 0
for r in data.get("test_results", []):
    row = id_to_row.get(r["test_id"])
    if not row:
        continue
    cls = r["classification"]
    ws.cell(row=row, column=8, value=cls)
    ws.cell(row=row, column=8).fill = fill_map.get(cls, err_fill)
    ws.cell(row=row, column=9, value=r.get("match_score", ""))
    ws.cell(row=row, column=10, value=r.get("keyword_score", ""))
    ws.cell(row=row, column=11, value=r.get("article_score", ""))
    ws.cell(row=row, column=12, value=", ".join(r.get("matched_articles", [])))
    updated += 1

wb.save(xlsx_file)
print(f"Updated {updated} rows in {xlsx_file}")
